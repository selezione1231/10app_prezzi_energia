import os
import re
import io
import zipfile
import requests
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from tqdm import tqdm

BASE_DIR = Path(__file__).resolve().parent
RAW_DIR = BASE_DIR / "data" / "raw"
PROCESSED_DIR = BASE_DIR / "data" / "processed"

RAW_DIR.mkdir(parents=True, exist_ok=True)
PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
}

def get_gme_download_url(year: int) -> str:
    return f"https://www.mercatoelettrico.org/it-it/Home/Esiti/Elettricita/MGP/Statistiche/DatiStorici/moduleId/10874/controller/GmeDatiStoriciItem/action/DownloadFile?fileName=Anno{year}.zip"

def download_year_zip(year: int) -> Path:
    zip_path = RAW_DIR / f"Anno{year}.zip"
    if zip_path.exists() and zip_path.stat().st_size > 1000:
        print(f"[CACHE] Anno {year} gia presente in cache ({zip_path.stat().st_size / 1e6:.2f} MB).")
        return zip_path

    url = get_gme_download_url(year)
    print(f"[DOWNLOAD] Scaricamento Anno {year} da {url}...")
    response = requests.get(url, headers=HEADERS, timeout=60)
    response.raise_for_status()

    with open(zip_path, "wb") as f:
        f.write(response.content)
    print(f"[OK] Anno {year} scaricato ({len(response.content) / 1e6:.2f} MB).")
    return zip_path

def parse_year_excel(zip_path: Path, year: int) -> pd.DataFrame:
    print(f"[PARSING] Elaborazione dati per anno {year}...")
    with zipfile.ZipFile(zip_path, "r") as z:
        namelist = z.namelist()
        
        # In recent years (2025+), there are _60 (hourly) and _15 (quarter-hourly) files. Prefer _60 for hourly data
        target_name = None
        for name in namelist:
            if "_60.xlsx" in name or "_60.xls" in name:
                target_name = name
                break
        if not target_name:
            for name in namelist:
                if name.lower().endswith((".xlsx", ".xls")):
                    target_name = name
                    break
        
        if not target_name:
            raise ValueError(f"Nessun file Excel trovato nello zip per l'anno {year}: {namelist}")

        file_bytes = z.read(target_name)

    # Determine sheet name
    if target_name.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        sheetnames = wb.sheetnames
        sheet_to_load = None
        for s in sheetnames:
            if "prezz" in s.lower() or "price" in s.lower():
                sheet_to_load = s
                break
        if not sheet_to_load and len(sheetnames) > 1:
            sheet_to_load = sheetnames[1]
        elif not sheet_to_load:
            sheet_to_load = 0
    else:
        import xlrd
        book = xlrd.open_workbook(file_contents=file_bytes, on_demand=True)
        sheetnames = book.sheet_names()
        sheet_to_load = None
        for s in sheetnames:
            if "prezz" in s.lower() or "price" in s.lower():
                sheet_to_load = s
                break
        if not sheet_to_load and len(sheetnames) > 1:
            sheet_to_load = sheetnames[1]
        elif not sheet_to_load:
            sheet_to_load = 0

    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_to_load)

    # Standardize column headers
    cleaned_cols = []
    for c in df.columns:
        c_str = str(c).strip().replace("\n", " ").replace("\r", "")
        if "Data" in c_str or "Date" in c_str:
            cleaned_cols.append("Data")
        elif "Ora" in c_str or "Hour" in c_str:
            cleaned_cols.append("Ora")
        else:
            cleaned_cols.append(c_str.upper())
    df.columns = cleaned_cols

    # Ensure Data and Ora exist
    if "Data" not in df.columns or "Ora" not in df.columns:
        raise KeyError(f"Colonne 'Data' o 'Ora' non trovate in {target_name}. Colonne trovate: {df.columns.tolist()}")

    # Clean date and hour
    df = df.dropna(subset=["Data", "Ora"]).copy()
    df["Data"] = df["Data"].astype(str).str.split(".").str[0].str.split(" ").str[0]
    df["Data"] = df["Data"].str.replace("-", "").str.replace("/", "")
    # Format YYYYMMDD
    df["Data"] = df["Data"].apply(lambda x: f"{int(float(x)):08d}" if str(x).isdigit() else str(x))
    
    df["Ora"] = pd.to_numeric(df["Ora"], errors="coerce").fillna(1).astype(int)

    # Create timestamp
    def make_timestamp(row):
        try:
            d_str = str(row["Data"])
            h = int(row["Ora"])
            return pd.to_datetime(d_str, format="%Y%m%d") + pd.Timedelta(hours=h-1)
        except Exception:
            return pd.NaT

    df["timestamp"] = df.apply(make_timestamp, axis=1)
    df = df.dropna(subset=["timestamp"])

    # Convert numeric price columns
    price_cols = [c for c in df.columns if c not in ["Data", "Ora", "timestamp"]]
    for col in price_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # Set timestamp as index and sort
    df = df.set_index("timestamp").sort_index()
    print(f"[OK] Anno {year}: {len(df)} record orari estratti.")
    return df

def build_dataset(start_year: int = 2015, end_year: int = 2026):
    print(f"=== INIZIO ACQUISIZIONE DATI STORICI GME ({start_year} - {end_year}) ===")
    dfs = []
    for y in range(start_year, end_year + 1):
        try:
            zip_p = download_year_zip(y)
            df_y = parse_year_excel(zip_p, y)
            dfs.append(df_y)
        except Exception as e:
            print(f"[ERRORE] Impossibile scaricare o elaborare l'anno {y}: {e}")

    if not dfs:
        raise RuntimeError("Nessun dato scaricato con successo.")

    print("\n[MERGE] Unione di tutte le annate...")
    full_df = pd.concat(dfs, axis=0)
    full_df = full_df[~full_df.index.duplicated(keep="last")].sort_index()

    main_zones = ["PUN", "NORD", "CNOR", "CSUD", "SUD", "SICI", "SARD", "CALA", "ROSN"]
    available_main_zones = [z for z in main_zones if z in full_df.columns]
    other_cols = [c for c in full_df.columns if c not in available_main_zones and c not in ["Data", "Ora"]]

    ordered_cols = ["Data", "Ora"] + available_main_zones + other_cols
    full_df = full_df[[c for c in ordered_cols if c in full_df.columns]]

    parquet_file = PROCESSED_DIR / f"gme_prezzi_zonali_{start_year}_{end_year}.parquet"
    csv_file = PROCESSED_DIR / f"gme_prezzi_zonali_{start_year}_{end_year}.csv"

    print(f"[EXPORT] Salvataggio Parquet: {parquet_file} ...")
    full_df.to_parquet(parquet_file)

    print(f"[EXPORT] Salvataggio CSV: {csv_file} ...")
    full_df.to_csv(csv_file)

    print("\n" + "="*60)
    print(f"Dataset storico completato con successo!")
    print(f"Intervallo temporale: {full_df.index.min()} -> {full_df.index.max()}")
    print(f"Totale ore registrate: {len(full_df):,} ore")
    print(f"Zone presenti: {available_main_zones}")
    print(f"Dimensione Parquet: {parquet_file.stat().st_size / 1e6:.2f} MB")
    print(f"Dimensione CSV: {csv_file.stat().st_size / 1e6:.2f} MB")
    print("="*60)

    return full_df

if __name__ == "__main__":
    df = build_dataset(start_year=2015, end_year=2026)
