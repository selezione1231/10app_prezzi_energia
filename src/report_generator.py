import io
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, Any

# Excel formatting with openpyxl
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF generation with reportlab
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def generate_excel_report(backtest_data: Dict[str, Any]) -> bytes:
    """Genera un report Excel completo e professionale con formattazione numerica e layout avanzato."""
    kpis = backtest_data["kpis"]
    quarterly_df = backtest_data["quarterly_df"]
    yearly_df = backtest_data["yearly_df"]
    monthly_df = backtest_data["monthly_df"]
    hourly_df = backtest_data["hourly_df"]

    wb = openpyxl.Workbook()
    # Rimuovi foglio default
    wb.remove(wb.active)

    # Stili
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    sub_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # ----------------------------------------------------
    # FOGLIO 1: SINTESI ESECUTIVA
    # ----------------------------------------------------
    ws_kpi = wb.create_sheet(title="Executive Summary")
    ws_kpi.views.sheetView[0].showGridLines = True

    ws_kpi["A1"] = "REPORT VALUTAZIONE ECONOMICA IMPIANTO FOTOVOLTAICO"
    ws_kpi["A1"].font = title_font
    ws_kpi["A2"] = f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')} | Mercato del Giorno Prima (GME MGP)"
    ws_kpi["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")

    ws_kpi["A4"] = "PARAMETRI IMPIANTO & SIMULAZIONE"
    ws_kpi["A4"].font = bold_font
    ws_kpi["A4"].fill = sub_header_fill

    params_rows = [
        ("Nome Impianto", kpis["plant_name"]),
        ("Potenza Nominale (kWp)", kpis["power_kw"]),
        ("Zona di Mercato Elettrico", kpis["zone"]),
        ("Data Inizio Backtest", kpis["start_date"]),
        ("Data Fine Backtest", kpis["end_date"]),
        ("Ore Totali Analizzate", kpis["tot_hours"]),
    ]
    for r_idx, (label, val) in enumerate(params_rows, start=5):
        ws_kpi.cell(row=r_idx, column=1, value=label).font = bold_font
        ws_kpi.cell(row=r_idx, column=2, value=val)

    ws_kpi["D4"] = "KPI ECONOMICI & PRESTAZIONALI"
    ws_kpi["D4"].font = bold_font
    ws_kpi["D4"].fill = sub_header_fill

    kpi_rows = [
        ("Produzione Totale (MWh)", kpis["tot_mwh"], "#,##0.00"),
        ("Ricavo Economico Totale (€)", kpis["tot_ricavo_eur"], "€ #,##0.00"),
        ("Prezzo Catturato Medio (€/MWh)", kpis["prezzo_catturato_eur_mwh"], "€ #,##0.00"),
        ("Prezzo Medio Zonale MGP (€/MWh)", kpis["prezzo_zonale_medio_eur_mwh"], "€ #,##0.00"),
        ("Capture Rate (%)", kpis["capture_rate_pct"] / 100.0, "0.0%"),
        ("Ricavo Specifico (€/kWp)", kpis["ricavo_specifico_eur_kwp"], "€ #,##0.00"),
        ("Ore Equivalenti Annue (h/anno)", kpis["ore_equivalenti_annue"], "#,##0.0"),
        ("Ore Produzione con Prezzo <= 0 €", kpis["ore_zero_prezzo_totali"], "#,##0")
    ]
    for r_idx, (label, val, fmt) in enumerate(kpi_rows, start=5):
        ws_kpi.cell(row=r_idx, column=4, value=label).font = bold_font
        c = ws_kpi.cell(row=r_idx, column=5, value=val)
        c.number_format = fmt
        c.font = bold_font

    # ----------------------------------------------------
    # FOGLIO 2: BREAKDOWN TRIMESTRALE (QUARTERLY)
    # ----------------------------------------------------
    ws_q = wb.create_sheet(title="Breakdown Trimestrale")
    ws_q.views.sheetView[0].showGridLines = True

    q_headers = [
        "Anno", "Trimestre", "Periodo", "Produzione (MWh)", "Ricavi (€)",
        "Prezzo Catturato (€/MWh)", "Prezzo Medio Zonale (€/MWh)", "Capture Rate (%)", "Ore Prezzo <= 0 €"
    ]
    for col_idx, h in enumerate(q_headers, 1):
        cell = ws_q.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in quarterly_df.iterrows():
        r = row_idx + 2
        ws_q.cell(row=r, column=1, value=row["anno"]).alignment = Alignment(horizontal="center")
        ws_q.cell(row=r, column=2, value=row["quarter"]).alignment = Alignment(horizontal="center")
        ws_q.cell(row=r, column=3, value=row["anno_trimestre"]).alignment = Alignment(horizontal="center")
        
        c_mwh = ws_q.cell(row=r, column=4, value=row["produzione_mwh"])
        c_mwh.number_format = "#,##0.00"
        
        c_ric = ws_q.cell(row=r, column=5, value=row["ricavo_eur"])
        c_ric.number_format = "€ #,##0.00"
        
        c_cap = ws_q.cell(row=r, column=6, value=row["prezzo_catturato_eur_mwh"])
        c_cap.number_format = "€ #,##0.00"
        
        c_zon = ws_q.cell(row=r, column=7, value=row["prezzo_medio_zonale_eur_mwh"])
        c_zon.number_format = "€ #,##0.00"
        
        c_rate = ws_q.cell(row=r, column=8, value=row["capture_rate_pct"] / 100.0)
        c_rate.number_format = "0.0%"
        
        c_zero = ws_q.cell(row=r, column=9, value=row["ore_zero_prezzo"])
        c_zero.number_format = "#,##0"

    # ----------------------------------------------------
    # FOGLIO 3: BREAKDOWN ANNUALE
    # ----------------------------------------------------
    ws_y = wb.create_sheet(title="Breakdown Annuale")
    ws_y.views.sheetView[0].showGridLines = True

    y_headers = [
        "Anno", "Produzione (MWh)", "Ricavi (€)", "Prezzo Catturato (€/MWh)",
        "Prezzo Medio Zonale (€/MWh)", "Capture Rate (%)", "Ore Prezzo <= 0 €"
    ]
    for col_idx, h in enumerate(y_headers, 1):
        cell = ws_y.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in yearly_df.iterrows():
        r = row_idx + 2
        ws_y.cell(row=r, column=1, value=row["anno"]).alignment = Alignment(horizontal="center")
        
        c_mwh = ws_y.cell(row=r, column=2, value=row["produzione_mwh"])
        c_mwh.number_format = "#,##0.00"
        
        c_ric = ws_y.cell(row=r, column=3, value=row["ricavo_eur"])
        c_ric.number_format = "€ #,##0.00"
        
        c_cap = ws_y.cell(row=r, column=4, value=row["prezzo_catturato_eur_mwh"])
        c_cap.number_format = "€ #,##0.00"
        
        c_zon = ws_y.cell(row=r, column=5, value=row["prezzo_medio_zonale_eur_mwh"])
        c_zon.number_format = "€ #,##0.00"
        
        c_rate = ws_y.cell(row=r, column=6, value=row["capture_rate_pct"] / 100.0)
        c_rate.number_format = "0.0%"
        
        c_zero = ws_y.cell(row=r, column=7, value=row["ore_zero_prezzo"])
        c_zero.number_format = "#,##0"

    # ----------------------------------------------------
    # FOGLIO 4: SERIE ORARIA (Prime 10.000 righe o completo)
    # ----------------------------------------------------
    ws_h = wb.create_sheet(title="Dati Orari Dettaglio")
    ws_h.views.sheetView[0].showGridLines = True

    h_headers = ["Timestamp", "Data", "Ora", "Produzione (kWh)", "Produzione (MWh)", "Prezzo Zonale (€/MWh)", "Ricavo (€)"]
    for col_idx, h in enumerate(h_headers, 1):
        cell = ws_h.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    sample_hourly = hourly_df.reset_index()
    for row_idx, row in sample_hourly.iterrows():
        r = row_idx + 2
        ws_h.cell(row=r, column=1, value=str(row["timestamp"]))
        ws_h.cell(row=r, column=2, value=f"{row['anno']}-{row['mese']:02d}-{row['giorno']:02d}")
        ws_h.cell(row=r, column=3, value=row["ora"])
        
        c_kwh = ws_h.cell(row=r, column=4, value=row["produzione_kwh"])
        c_kwh.number_format = "#,##0.00"
        
        c_mwh = ws_h.cell(row=r, column=5, value=row["produzione_mwh"])
        c_mwh.number_format = "#,##0.0000"
        
        c_pr = ws_h.cell(row=r, column=6, value=row["prezzo_zonale_eur_mwh"])
        c_pr.number_format = "€ #,##0.00"
        
        c_ric = ws_h.cell(row=r, column=7, value=row["ricavo_eur"])
        c_ric.number_format = "€ #,##0.00"

    # Auto-adatta larghezza colonne per tutti i fogli
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_pdf_report(backtest_data: Dict[str, Any]) -> bytes:
    """Genera un report PDF professionale con tabelle di sintesi, trimestrali e annuali."""
    kpis = backtest_data["kpis"]
    quarterly_df = backtest_data["quarterly_df"]
    yearly_df = backtest_data["yearly_df"]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm
    )

    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#555555'),
        spaceAfter=15
    )

    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontSize=13,
        leading=16,
        textColor=colors.HexColor('#1F4E78'),
        spaceBefore=12,
        spaceAfter=8
    )

    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10)
    cell_bold = ParagraphStyle('CellBold', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica-Bold')
    cell_center = ParagraphStyle('CellCenter', parent=styles['Normal'], fontSize=8, leading=10, alignment=1)

    story = []

    # Intestazione
    story.append(Paragraph("REPORT ECONOMICO PRODUZIONE FOTOVOLTAICA", title_style))
    story.append(Paragraph(f"Simulazione e Backtest su Mercato Zonale Italiano (MGP) | {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1F4E78'), spaceAfter=15))

    # Tabella Parametri e KPI Principali (2 colonne)
    summary_data = [
        [
            Paragraph("<b>Parametro Impianto</b>", cell_bold), Paragraph("<b>Valore</b>", cell_bold),
            Paragraph("<b>Indicatore Economico</b>", cell_bold), Paragraph("<b>Risultato</b>", cell_bold)
        ],
        [
            Paragraph("Impianto", cell_style), Paragraph(str(kpis["plant_name"]), cell_style),
            Paragraph("Produzione Totale", cell_style), Paragraph(f"<b>{kpis['tot_mwh']:,} MWh</b>", cell_style)
        ],
        [
            Paragraph("Potenza Installata", cell_style), Paragraph(f"<b>{kpis['power_kw']:,} kWp</b>", cell_style),
            Paragraph("Ricavi Totali Generati", cell_style), Paragraph(f"<font color='#1F4E78'><b>€ {kpis['tot_ricavo_eur']:,.2f}</b></font>", cell_style)
        ],
        [
            Paragraph("Zona di Mercato", cell_style), Paragraph(f"<b>{kpis['zone']}</b>", cell_style),
            Paragraph("Prezzo Catturato Medio", cell_style), Paragraph(f"<b>€ {kpis['prezzo_catturato_eur_mwh']:.2f} /MWh</b>", cell_style)
        ],
        [
            Paragraph("Periodo Analizzato", cell_style), Paragraph(f"{kpis['start_date']} -> {kpis['end_date']}", cell_style),
            Paragraph("Prezzo Medio Zonale MGP", cell_style), Paragraph(f"€ {kpis['prezzo_zonale_medio_eur_mwh']:.2f} /MWh", cell_style)
        ],
        [
            Paragraph("Ore Totali di Analisi", cell_style), Paragraph(f"{kpis['tot_hours']:,} ore", cell_style),
            Paragraph("Capture Rate Solare", cell_style), Paragraph(f"<b>{kpis['capture_rate_pct']:.1f}%</b>", cell_style)
        ],
        [
            Paragraph("Ore Prezzo <= 0 €", cell_style), Paragraph(f"{kpis['ore_zero_prezzo_totali']} ore", cell_style),
            Paragraph("Ricavo Specifico", cell_style), Paragraph(f"<b>€ {kpis['ricavo_specifico_eur_kwp']:.2f} /kWp</b>", cell_style)
        ]
    ]

    t_summary = Table(summary_data, colWidths=[3.5*cm, 4.5*cm, 4.5*cm, 4.5*cm])
    t_summary.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t_summary)
    story.append(Spacer(1, 15))

    # Tabella Annuale
    story.append(Paragraph("1. Sintesi Annuale di Produzione e Ricavi", section_style))
    y_table_data = [[
        Paragraph("<b>Anno</b>", cell_center),
        Paragraph("<b>Produzione (MWh)</b>", cell_center),
        Paragraph("<b>Ricavi Totali (€)</b>", cell_center),
        Paragraph("<b>Prezzo Catturato (€/MWh)</b>", cell_center),
        Paragraph("<b>Prezzo Zonale (€/MWh)</b>", cell_center),
        Paragraph("<b>Capture Rate</b>", cell_center),
        Paragraph("<b>Ore P <= 0€</b>", cell_center)
    ]]

    for _, row in yearly_df.iterrows():
        y_table_data.append([
            Paragraph(str(row["anno"]), cell_center),
            Paragraph(f"{row['produzione_mwh']:,.2f}", cell_center),
            Paragraph(f"€ {row['ricavo_eur']:,.2f}", cell_center),
            Paragraph(f"€ {row['prezzo_catturato_eur_mwh']:.2f}", cell_center),
            Paragraph(f"€ {row['prezzo_medio_zonale_eur_mwh']:.2f}", cell_center),
            Paragraph(f"{row['capture_rate_pct']:.1f}%", cell_center),
            Paragraph(str(int(row["ore_zero_prezzo"])), cell_center)
        ])

    t_yearly = Table(y_table_data, colWidths=[2*cm, 2.8*cm, 3.2*cm, 3.2*cm, 3*cm, 2.3*cm, 1.8*cm])
    t_yearly.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t_yearly)
    story.append(Spacer(1, 15))

    # Tabella Trimestrale
    story.append(Paragraph("2. Dettaglio Trimestrale (Q1, Q2, Q3, Q4)", section_style))
    q_table_data = [[
        Paragraph("<b>Periodo</b>", cell_center),
        Paragraph("<b>Trimestre</b>", cell_center),
        Paragraph("<b>Produzione (MWh)</b>", cell_center),
        Paragraph("<b>Ricavi (€)</b>", cell_center),
        Paragraph("<b>Prezzo Catturato (€/MWh)</b>", cell_center),
        Paragraph("<b>Prezzo Zonale (€/MWh)</b>", cell_center),
        Paragraph("<b>Capture Rate</b>", cell_center)
    ]]

    for _, row in quarterly_df.iterrows():
        q_table_data.append([
            Paragraph(str(row["anno_trimestre"]), cell_center),
            Paragraph(str(row["quarter"]), cell_center),
            Paragraph(f"{row['produzione_mwh']:,.2f}", cell_center),
            Paragraph(f"€ {row['ricavo_eur']:,.2f}", cell_center),
            Paragraph(f"€ {row['prezzo_catturato_eur_mwh']:.2f}", cell_center),
            Paragraph(f"€ {row['prezzo_medio_zonale_eur_mwh']:.2f}", cell_center),
            Paragraph(f"{row['capture_rate_pct']:.1f}%", cell_center)
        ])

    t_quarterly = Table(q_table_data, colWidths=[2.5*cm, 2*cm, 2.8*cm, 3.2*cm, 3.2*cm, 2.8*cm, 2*cm])
    t_quarterly.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(t_quarterly)

    # Note conclusive
    story.append(Spacer(1, 20))
    story.append(Paragraph("<b>Note Metodologiche:</b><br/>"
                           "• Il Prezzo Catturato (Capture Price) rappresenta la media ponderata dei prezzi orari sulla produzione solare effettiva (Σ(Produzione × Prezzo) / Σ Produzione).<br/>"
                           "• Il Capture Rate misura lo sconto/premio rispetto al prezzo medio aritmetico della zona per via del profilo di produzione solare (duck curve).<br/>"
                           "• Dati di mercato orari ufficiali GME MGP 2015-2026.",
                           subtitle_style))

    doc.build(story)
    return buffer.getvalue()
